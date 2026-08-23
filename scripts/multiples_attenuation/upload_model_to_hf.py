"""
Upload trained model checkpoints (best.pt + config.yaml) to a Hugging Face repository.

Usage:
    export HF_NAMESPACE=GeoBrain  # or HF_USERNAME (personal account)
    export HF_TOKEN="your_hf_token"
    python scripts/multiples_attenuation/upload_model_to_hf.py --models unet res_unet

Choose from: unet, unet_plus, res_unet, res_unet_plus, dncnn, atten_unet,
              atten_unet_plus, kiraz_trace_cnn, sagan, dnndat

Optional:
    --repo-name NAME     HF repo name (default: multiples-attenuation)
    --results-dir PATH   Override results root (default: /data/shared/benchmark/multiples/results)
    --dry-run            Scan and print what would be uploaded without uploading
    --models M [M ...]   Only upload the listed model(s); default = all
"""

import argparse
import logging
import os
import re
import sys

from huggingface_hub import HfApi, create_repo

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

RESULTS_ROOT = "/data/shared/benchmark/multiples/results"
DEFAULT_REPO = "multiples-attenuation"

FOLDER_PATTERN = re.compile(
    r"denoise_(?P<model>.+)_base\d+_level(?P<level>[\d.]+)_seed(?P<seed>\d+)"
)

# Fixed-dataset format: denoise_{model}_base{date}_seed{seed}
FOLDER_PATTERN_SEED_ONLY = re.compile(
    r"denoise_(?P<model>.+)_base\d+_seed(?P<seed>\d+)$"
)

# Simplified format (new training scripts): denoise_{model}_base{date}
FOLDER_PATTERN_SIMPLE = re.compile(
    r"denoise_(?P<model>.+)_base\d+$"
)

# Extract noise level from data path, e.g. noisy_1.0.sgy → 1.0
_NOISE_LEVEL_RE = re.compile(r"noisy_([\d.]+)\.sgy")

MODEL_DISPLAY = {
    "unet": "UNet",
    "unet_plus": "UNet-Plus",
    "res_unet": "ResUNet",
    "res_unet_plus": "ResUNet-Plus",
    "dncnn": "DnCNN",
    "atten_unet": "Attention UNet",
    "atten_unet_plus": "Attention UNet-Plus",
    "kiraz_trace_cnn": "Kiraz Trace CNN",
    "sagan": "SAGAN",
    "dnndat": "DNNDAT",
}

MODEL_DESCRIPTION = {
    "unet": "Classic encoder-decoder with skip connections (Ronneberger et al., 2015). Base channels: 32, depth: 4.",
    "unet_plus": "Wider UNet variant. Base channels: 64, depth: 4.",
    "res_unet": "U-Net with residual blocks replacing plain double-conv layers (He et al., 2016; Zhang et al., 2018). Base channels: 32, depth: 4.",
    "res_unet_plus": "Wider ResUNet variant. Base channels: 64, depth: 4.",
    "dncnn": "Flat 17-layer Conv-BN-ReLU stack with residual learning (Zhang et al., 2017, IEEE TIP). Base channels: 64.",
    "atten_unet": "U-Net with additive attention gates on skip connections (Oktay et al., 2018, MIDL). Base channels: 32, depth: 4.",
    "atten_unet_plus": "Wider Attention UNet variant. Base channels: 64, depth: 4.",
    "kiraz_trace_cnn": "Trace-by-trace 1D CNN for free-surface multiple attenuation and deghosting (Kiraz et al., 2024). Nine hidden Conv1d layers with large temporal kernels.",
    "sagan": "Self-attention GAN generator for seismic surface-related multiple suppression (Tao et al., 2022). U-Net generator with a bottleneck self-attention block.",
    "dnndat": "DNNDAT-style convolutional encoder-decoder for marine multiple suppression (Wang et al., 2022). U-Net-like encoder-decoder with 28 convolutional layers and dropout.",
}

# User-facing name → internal folder-pattern key
_MODEL_ALIASES = {
    "attention_unet": "atten_unet",
    "attention_unet_plus": "atten_unet_plus",
    "kiraz": "kiraz_trace_cnn",
}

_ALL_MODELS = sorted(
    ["unet", "unet_plus", "res_unet", "res_unet_plus", "dncnn",
     "atten_unet", "atten_unet_plus", "kiraz_trace_cnn", "sagan", "dnndat"]
)

_VALID_MODEL_NAMES = sorted(list(_MODEL_ALIASES.keys()) + _ALL_MODELS)


def _resolve_models(requested):
    """Resolve user-facing model names to internal folder-pattern keys."""
    if not requested:
        return _ALL_MODELS
    resolved = []
    for name in requested:
        name = name.strip().lower()
        if name in _MODEL_ALIASES:
            name = _MODEL_ALIASES[name]
        if name not in _ALL_MODELS:
            raise ValueError(
                f"Unknown model {name!r}. Valid choices: {', '.join(_VALID_MODEL_NAMES)}"
            )
        resolved.append(name)
    return resolved

# Default path to the batch evaluation Excel file
BATCH_EVAL_XLSX = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "multiples_attenuation", "batch_evaluation.xlsx",
)


def parse_batch_eval_xlsx(xlsx_path: str):
    """Parse batch evaluation Excel into structured data.

    Returns
    -------
    dict : {level_str: {model_key: {metric: value_str}}}.
        model_key "raw" maps to the noisy baseline.
    """
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl not installed; cannot parse batch evaluation Excel.")
        return {}

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    result = {}
    for sheet_name in wb.sheetnames:
        # sheet name format: "Multiples" for the fixed dataset, or "Noise X.X".
        level = "single" if sheet_name == "Multiples" else sheet_name.replace("Noise ", "").strip()
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        headers = [str(h).strip() if h else "" for h in rows[0]]
        # map display names to registry keys
        display_to_key = {v: k for k, v in MODEL_DISPLAY.items()}
        display_to_key["Raw (noisy)"] = "raw"

        level_data = {}
        for row in rows[1:]:
            if not row or not row[0]:
                continue
            method_name = str(row[0]).strip()
            model_key = display_to_key.get(method_name, method_name.lower().replace(" ", "_"))
            metrics = {}
            for ci in range(1, len(headers)):
                if ci < len(row) and row[ci] is not None:
                    metrics[headers[ci].lower()] = str(row[ci]).strip()
            if metrics:
                level_data[model_key] = metrics
        if level_data:
            result[level] = level_data
    wb.close()
    return result


def generate_model_card(entries, repo_name: str, eval_data: dict = None) -> str:
    """Generate a Hugging Face model card README.md."""
    model_keys = sorted({e["model"] for e in entries})
    models_str = "\n".join(
        f"- **{MODEL_DISPLAY.get(k, k)}** (`{k}`) — {MODEL_DESCRIPTION.get(k, '')}"
        for k in model_keys
    )

    count_str = "\n".join(
        f"  - {MODEL_DISPLAY.get(k, k)}: {sum(1 for e in entries if e['model'] == k)} checkpoints"
        for k in model_keys
    )

    # build results tables from batch evaluation data
    results_section = _build_results_section(eval_data or {})

    card = f"""---
tags:
- seismic
- multiples
- denoising
- unet
- resunet
- dncnn
- attention-unet
- pytorch
library_name: pytorch
---

# Marine Multiples Attenuation Benchmark

Deep-learning-based free-surface multiple attenuation on marine pre-stack seismic shot gathers.

## Task

Given a shot gather containing primaries and free-surface multiples, the model predicts the additive multiple component. The attenuated primary estimate is obtained by:

```
denoised = noisy_input - predicted_noise
```

This is a **paired regression** task trained with a noise-label objective. The supervision target is the multiple wavefield stored in the paired noise-label volume.

## Dataset

- **Input volume**: `/data/shared/benchmark/multiples/noisy/total_nodw.sgy`
- **Multiple-label volume**: `/data/shared/benchmark/multiples/noise/multiples.sgy`
- **Geometry**: regular shot gathers with 638 traces per shot
- **Split**: shot-level sequential split from the training configs, typically 510 train shots, 64 validation shots, and 64 held-out test shots

The uploaded checkpoints are trained on the fixed paired marine multiples benchmark used by `scripts/multiples_attenuation`.

Metrics are computed on the held-out test shots in the normalized domain. If a batch-evaluation workbook is supplied, its values are embedded below.

## Model Architectures

{models_str}

## Uploaded Checkpoints

{count_str}

### Preprocessing

- **Normalization**: `max_abs`, global scope — the entire dataset scaled to [-1, 1]
- **Patching**: overlapping 2D patches, usually 256 traces × 512 time samples with 50% overlap
- **Tensor format**: PyTorch NCHW patches `(batch, 1, trace, time)`

## Repository Structure

```
models/
├── unet/
│   ├── seed42/
│   │   ├── best.pt          # Best checkpoint (minimum validation loss)
│   │   └── config.yaml      # Full training configuration
│   ├── seed43/
│   ├── seed44/
│   └── ...
└── res_unet/
    └── ...
```

Each subdirectory corresponds to one experiment: a model architecture trained with a specific random seed.

## Training Details

| Hyperparameter | Value |
|----------------|-------|
| Loss | MSE on the predicted multiple/noise component |
| Optimizer | Adam / AdamW / SGD, depending on model config |
| Scheduler | Cosine annealing (min_lr=1e-6) |
| Epochs | 100-200, depending on model config |
| Gradient clipping | 1.0 (max norm) |
| Seeds | 42, 43, 44 per experiment |

## Usage

```python
import torch
from huggingface_hub import hf_hub_download

# Download a checkpoint
repo = "{repo_name}"
model_key = "res_unet"
seed = "42"

ckpt_path = hf_hub_download(
    repo_id=repo,
    filename=f"models/{{model_key}}/seed{{seed}}/best.pt",
)

# Load state dict
state_dict = torch.load(ckpt_path, map_location="cpu", weights_only=True)

# For full model loading, instantiate the corresponding architecture
# and load the state dict (see config.yaml for exact architecture params).
```

See the companion benchmark documentation for detailed experimental setup and full evaluation results.

{results_section}

## References

- Ronneberger et al., U-Net: Convolutional Networks for Biomedical Image Segmentation, MICCAI 2015
- He et al., Deep Residual Learning for Image Recognition, CVPR 2016
- Zhang et al., Image Denoising via Deep CNN (DnCNN), IEEE TIP 2017
- Oktay et al., Attention U-Net: Learning Where to Look for the Pancreas, MIDL 2018
- Kiraz et al., Attenuating free-surface multiples and ghost reflection from seismic data using a trace-by-trace convolutional neural network approach, Geophysical Prospecting 2024
- Tao et al., Seismic Surface-Related Multiples Suppression Based on SAGAN, IEEE Geoscience and Remote Sensing Letters 2022
- Wang et al., Seismic multiple suppression based on a deep neural network method for marine data, Geophysics 2022
"""
    return card


def _build_results_section(eval_data: dict) -> str:
    """Build markdown results tables from batch evaluation data."""
    if not eval_data:
        return """## Results

*Results pending — run batch_evaluate.py to populate.*
"""

    # column order
    metric_cols = ["parameters (m)", "snr", "psnr", "ssim", "mae", "mse", "rmse"]
    metric_headers = ["Method", "Params (M)", "SNR (dB)", "PSNR (dB)", "SSIM", "MAE", "MSE", "RMSE"]

    row_order = [
        "raw",
        "unet",
        "unet_plus",
        "res_unet",
        "res_unet_plus",
        "dncnn",
        "atten_unet",
        "atten_unet_plus",
        "kiraz_trace_cnn",
        "sagan",
        "dnndat",
    ]

    lines = ["## Results\n"]
    lines.append("Mean ± std over available seeds on the held-out test shots, evaluated on 2D-flattened data in the normalized domain. Raw (noisy) is the input before multiple attenuation.\n")

    def _level_key(level: str):
        if level == "single":
            return 0, 0.0
        try:
            return 1, float(level)
        except ValueError:
            return 2, level

    sorted_levels = sorted(eval_data.keys(), key=_level_key)

    for level in sorted_levels:
        title = "Multiples" if level == "single" else f"Noise Level {level}"
        lines.append(f"### {title}\n")
        # header
        lines.append("| " + " | ".join(metric_headers) + " |")
        # separator
        sep = ["---"] + [":---:"] * (len(metric_headers) - 1)
        lines.append("| " + " | ".join(sep) + " |")

        level_data = eval_data[level]
        for rk in row_order:
            if rk not in level_data:
                continue
            metrics = level_data[rk]
            display = "Raw (noisy)" if rk == "raw" else MODEL_DISPLAY.get(rk, rk)
            row = [display]
            for mc in metric_cols:
                row.append(metrics.get(mc, "—"))
            lines.append("| " + " | ".join(row) + " |")
        lines.append("")

    return "\n".join(lines)


def scan_results(results_dir: str):
    """Scan results directory and return list of (folder_name, model, level, seed).

    Supports three naming formats:

    - Full:   ``denoise_{model}_base{date}_level{level}_seed{seed}``
    - Fixed:  ``denoise_{model}_base{date}_seed{seed}``
    - Simple: ``denoise_{model}_base{date}``  (reads seed/level from config.yaml)
    """
    import yaml as _yaml

    entries = []
    for name in os.listdir(results_dir):
        fpath = os.path.join(results_dir, name)
        if not os.path.isdir(fpath):
            continue

        m = FOLDER_PATTERN.match(name)
        if m:
            model, level, seed = m.group("model"), m.group("level"), m.group("seed")
        else:
            m = FOLDER_PATTERN_SEED_ONLY.match(name)
            if m:
                model, seed = m.group("model"), m.group("seed")
                level = "single"
            else:
                m = FOLDER_PATTERN_SIMPLE.match(name)
                if not m:
                    continue
                model = m.group("model")
                seed = "0"
                level = "single"
                config_path = os.path.join(fpath, "config.yaml")
                if os.path.isfile(config_path):
                    try:
                        with open(config_path, "r") as f:
                            cfg = _yaml.safe_load(f)
                        seed = str(cfg.get("experiment", {}).get("seed", "0"))
                        input_path = cfg.get("data", {}).get("segy_pair", {}).get("input_path", "")
                        lm = _NOISE_LEVEL_RE.search(input_path)
                        if lm:
                            level = lm.group(1)
                    except Exception:
                        pass

        best_pt = os.path.join(fpath, "checkpoints", "best.pt")
        config_yaml = os.path.join(fpath, "config.yaml")
        entries.append(
            {
                "folder": name,
                "path": fpath,
                "model": model,
                "level": level,
                "seed": seed,
                "best_pt": best_pt if os.path.isfile(best_pt) else None,
                "config_yaml": config_yaml if os.path.isfile(config_yaml) else None,
            }
        )
    entries.sort(key=lambda x: (
        x["model"],
        float(x["level"]) if x["level"].replace(".", "", 1).isdigit() else 0.0,
        int(x["seed"]) if x["seed"].lstrip("-").isdigit() else 0,
    ))
    return entries


def hf_path(entry: dict, filename: str) -> str:
    """Construct the path inside the HF repo."""
    if entry["level"] in ("single", "unknown", "", None):
        run_name = f"seed{entry['seed']}"
    else:
        run_name = f"level{entry['level']}_seed{entry['seed']}"
    subdir = f"{entry['model']}/{run_name}"
    return f"models/{subdir}/{filename}"


def main():
    parser = argparse.ArgumentParser(description="Upload model checkpoints to Hugging Face.")
    parser.add_argument("--repo-name", default=DEFAULT_REPO, help=f"HF repo name (default: {DEFAULT_REPO})")
    parser.add_argument(
        "--results-dir", default=RESULTS_ROOT, help=f"Results root (default: {RESULTS_ROOT})"
    )
    parser.add_argument("--dry-run", action="store_true", help="Scan and print what would be uploaded")
    parser.add_argument(
        "--no-model-card", action="store_true", help="Skip uploading the model card"
    )
    parser.add_argument(
        "--models", nargs="*", metavar="MODEL", default=None,
        help="Only upload the listed model(s).  Valid: %s.  Default: all."
             % ", ".join(_VALID_MODEL_NAMES),
    )
    parser.add_argument(
        "--eval-xlsx", nargs="+", default=None,
        help="Path(s) to batch evaluation Excel(s) for populating the model card.  "
             "When multiple files are given, their data is merged (later files "
             "override earlier ones for the same model at the same data level).  "
             "Default: scripts/multiples_attenuation/batch_evaluation.xlsx",
    )
    args = parser.parse_args()

    namespace = os.environ.get("HF_NAMESPACE") or os.environ.get("HF_USERNAME")
    token = os.environ.get("HF_TOKEN")

    scanned_entries = scan_results(args.results_dir)
    entries = scanned_entries

    # Filter to requested models (if --models given)
    selected = _resolve_models(args.models)
    if args.models:
        entries = [e for e in entries if e["model"] in selected]
        if not entries:
            logger.warning(
                "No result folders matched models: %s.  Available in scan: %s",
                ", ".join(selected),
                ", ".join(sorted({e["model"] for e in scanned_entries})),
            )
            sys.exit(0)
        logger.info(
            "Selected models: %s → %d folder(s)", ", ".join(selected), len(entries)
        )

    if not entries:
        logger.warning("No matching result folders found in %s", args.results_dir)
        sys.exit(0)

    missing_files = [
        e for e in entries if e["best_pt"] is None and e["config_yaml"] is None
    ]
    if missing_files:
        logger.info(
            "Skipping %d folder(s) with neither checkpoints/best.pt nor config.yaml.",
            len(missing_files),
        )
    entries = [e for e in entries if e["best_pt"] is not None or e["config_yaml"] is not None]
    if not entries:
        logger.warning("No uploadable checkpoint/config files found in %s", args.results_dir)
        sys.exit(0)

    repo_id = f"{namespace}/{args.repo_name}" if namespace else args.repo_name
    total = len(entries)
    found_pt = sum(1 for e in entries if e["best_pt"])
    found_yaml = sum(1 for e in entries if e["config_yaml"])

    logger.info("Found %d uploadable experiment folder(s):", total)
    logger.info("  - %d have checkpoints/best.pt", found_pt)
    logger.info("  - %d have config.yaml", found_yaml)

    if args.dry_run:
        logger.info("Dry-run mode — files that would be uploaded:")
        for e in entries:
            if e["best_pt"]:
                logger.info("  [upload] %s", hf_path(e, "best.pt"))
            if e["config_yaml"]:
                logger.info("  [upload] %s", hf_path(e, "config.yaml"))
        if not args.no_model_card:
            logger.info("  [upload] README.md (model card)")
        return

    if not token:
        logger.error("HF_TOKEN environment variable is not set.")
        sys.exit(1)

    api = HfApi()
    logger.info("Creating / ensuring repo: %s", repo_id)
    create_repo(repo_id, token=token, exist_ok=True, private=False)

    # Fetch existing files in the repo for incremental (skip-already-existing) upload
    try:
        existing = set(api.list_repo_files(repo_id, token=token))
        logger.info("Repo has %d existing file(s); will skip those.", len(existing))
    except Exception:
        logger.info("Could not list repo files (repo may be empty); uploading all.")
        existing = set()

    uploaded = 0
    skipped = 0
    failed = 0

    # Upload model card (always overwrite — it is regenerated each run)
    if not args.no_model_card:
        # Parse batch evaluation results for the model card
        eval_data: dict = {}
        xlsx_paths = args.eval_xlsx if args.eval_xlsx else [BATCH_EVAL_XLSX]
        for xlsx_path in xlsx_paths:
            if os.path.isfile(xlsx_path):
                data = parse_batch_eval_xlsx(xlsx_path)
                if data:
                    # merge: later files override earlier ones for same (level, model)
                    for level, models in data.items():
                        eval_data.setdefault(level, {}).update(models)
                    logger.info("Loaded batch evaluation data from %s", xlsx_path)
            else:
                logger.warning("Eval Excel not found: %s", xlsx_path)
        if not eval_data:
            logger.warning("No evaluation data loaded — model card will have "
                           "'Results pending' placeholder.")
        card = generate_model_card(entries, repo_id, eval_data)
        try:
            api.upload_file(
                path_or_fileobj=card.encode(),
                path_in_repo="README.md",
                repo_id=repo_id,
                token=token,
            )
            logger.info("  [OK]     README.md (model card)")
            uploaded += 1
        except Exception as exc:
            logger.error("  [FAIL]   README.md — %s", exc)
            failed += 1

    # Upload checkpoints and configs
    for e in entries:
        files_to_upload = {}
        if e["best_pt"]:
            files_to_upload["best.pt"] = e["best_pt"]
        if e["config_yaml"]:
            files_to_upload["config.yaml"] = e["config_yaml"]

        for fname, local_path in files_to_upload.items():
            remote = hf_path(e, fname)
            if remote in existing:
                logger.info("  [SKIP]   %s (already exists)", remote)
                skipped += 1
                continue
            try:
                with open(local_path, "rb") as f:
                    api.upload_file(
                        path_or_fileobj=f,
                        path_in_repo=remote,
                        repo_id=repo_id,
                        token=token,
                    )
                logger.info("  [OK]     %s", remote)
                uploaded += 1
            except Exception as exc:
                logger.error("  [FAIL]   %s — %s", remote, exc)
                failed += 1

    logger.info("Done. %d uploaded, %d skipped, %d error(s).", uploaded, skipped, failed)


if __name__ == "__main__":
    main()
