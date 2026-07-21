"""Batch evaluation: iterate experiment directories, run inference on held-out
test sets, compute metrics before/after denoising, and output an Excel workbook.

Usage::

    # Evaluate all experiments under a directory
    python scripts/multiples_attenuation/batch_evaluate.py \\
        --root_dir results/multiples_attenuation \\
        --output results/multiples_attenuation/batch_evaluation.xlsx

    # Only evaluate specific models
    python scripts/multiples_attenuation/batch_evaluate.py \\
        --root_dir /data/shared/benchmark/multiples/results \\
        --models unet res_unet dncnn atten_unet kiraz_trace_cnn sagan dnndat \\
        --output results/batch_evaluation_part.xlsx

    # Merge into existing Excel without re-evaluating already-done models
    python scripts/multiples_attenuation/batch_evaluate.py \\
        --root_dir /data/shared/benchmark/multiples/results \\
        --models atten_unet \\
        --output results/batch_evaluation.xlsx \\
        --merge

    # Specify device and batch size
    python scripts/multiples_attenuation/batch_evaluate.py \\
        --root_dir /data/shared/benchmark/multiples/results \\
        --models res_unet \\
        --device cuda:0 \\
        --batch_size 4
"""

from __future__ import annotations

import argparse
import re
import sys
import traceback
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import torch
import yaml

# ---------------------------------------------------------------------------
# repo-root bootstrap (same pattern as the training scripts)
# ---------------------------------------------------------------------------
_REPO_ROOT = next(
    (
        p
        for p in Path(__file__).resolve().parents
        if (p / "model").is_dir() and (p / "utils").is_dir()
    ),
    None,
)
if _REPO_ROOT is None:
    raise RuntimeError(
        "Cannot find repo root (a directory containing both ``model/`` and ``utils/``)."
    )
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from model.multiples_attenuation import build_model  # noqa: E402
from utils.inference_utils import (  # noqa: E402
    compute_binned_metrics,
    compute_shot_metrics,
    inference_on_shots,
    save_shot_visualizations,
    select_random_shots,
)

# ---------------------------------------------------------------------------
# constants
# ---------------------------------------------------------------------------
METRIC_NAMES = ["snr", "psnr", "ssim", "mae", "mse", "rmse"]
METRIC_DISPLAY = [m.upper() for m in METRIC_NAMES]  # SNR, PSNR, SSIM, MAE, MSE, RMSE

# registry key → display name (rows, top-to-bottom in each sheet)
MODEL_DISPLAY = {
    "unet": "UNet",
    "unet_plus": "UNet-Plus",
    "res_unet": "ResUNet",
    "res_unet_plus": "ResUNet-Plus",
    "dncnn": "DnCNN",
    "atten_unet": "Attention UNet",
    "atten_unet_plus": "Attention UNet-Plus",
    "kiraz_trace_cnn": "Kiraz Trace CNN",
    "sagan": "SAGAN",
    "dnndat": "DNNDAT",
}
MODEL_ROW_ORDER = [
    "unet", "unet_plus",
    "res_unet", "res_unet_plus",
    "dncnn",
    "atten_unet", "atten_unet_plus",
    "kiraz_trace_cnn",
    "sagan",
    "dnndat",
]


# directory-name pattern: denoise_{model}_base{date}_level{level}_seed{seed}
_DIR_RE = re.compile(
    r"^denoise_(.+)_base\d+_level([\d.]+)_seed(\d+)$"
)

# Fixed-dataset format: denoise_{model}_base{date}_seed{seed}
_DIR_RE_SEED_ONLY = re.compile(r"^denoise_(.+)_base\d+_seed(\d+)$")

# Simplified format (new training scripts): denoise_{model}_base{date}
_DIR_RE_SIMPLE = re.compile(r"^denoise_(.+)_base(\d+)$")

# Extract noise level from data path, e.g. noisy_1.0.sgy → 1.0
_NOISE_LEVEL_RE = re.compile(r"noisy_([\d.]+)\.sgy")


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------
def parse_dir_name(name: str) -> Optional[Tuple[str, str, str]]:
    """Return ``(model_name, data_level, seed)`` or ``None`` if unparseable."""
    m = _DIR_RE.match(name)
    if m is not None:
        return m.group(1), m.group(2), m.group(3)
    m = _DIR_RE_SEED_ONLY.match(name)
    if m is not None:
        return m.group(1), "single", m.group(2)
    return None


def _parse_noise_level_from_dir(result_dir: Path) -> str:
    """Extract data level from a result directory name."""
    parsed = parse_dir_name(result_dir.name)
    if parsed is not None:
        return parsed[1]
    m = _DIR_RE_SIMPLE.match(result_dir.name)
    if m is not None:
        config_path = result_dir / "config.yaml"
        if config_path.is_file():
            try:
                with open(config_path, "r") as f:
                    cfg = yaml.safe_load(f)
                input_path = cfg.get("data", {}).get("segy_pair", {}).get("input_path", "")
                lm = _NOISE_LEVEL_RE.search(input_path)
                if lm:
                    return lm.group(1)
            except Exception:
                pass
    return "unknown"


def _level_sort_key(level: str) -> Tuple[int, float | str]:
    """Sort numeric noise levels first and fixed-dataset labels last."""
    try:
        return 0, float(level)
    except (TypeError, ValueError):
        return 1, str(level)


def _sheet_title(level: str) -> str:
    """Return a valid Excel sheet title for a data level."""
    if level == "single":
        return "Multiples"
    return f"Noise {level}"


def _save_viz_and_npy(
    eval_dir: Path,
    input_shots: np.ndarray,
    clean_shots: np.ndarray,
    denoised_shots: np.ndarray,
    model_type: str,
    data_level: str,
    n_viz_shots: int = 5,
) -> None:
    """Save full-volume .npy files and per-shot visualizations to ``eval_dir``.

    Directory layout (aligned with ``inference_denoise_*.py``)::

        eval_dir/
        ├── npy/
        │   ├── input_shots.npy
        │   ├── target_shots.npy
        │   └── pred_shots.npy
        └── visualizations/
            ├── denoise_{model}_level{level}_shot_0000.png
            ├── …
            └── data/
                ├── denoise_{model}_level{level}_shot_0000_input.npy
                ├── denoise_{model}_level{level}_shot_0000_prediction.npy
                ├── denoise_{model}_level{level}_shot_0000_target.npy
                └── …
    """
    eval_dir.mkdir(parents=True, exist_ok=True)

    # --- full-volume .npy files ---
    npy_dir = eval_dir / "npy"
    npy_dir.mkdir(parents=True, exist_ok=True)
    np.save(npy_dir / "input_shots.npy", input_shots)
    np.save(npy_dir / "target_shots.npy", clean_shots)
    np.save(npy_dir / "pred_shots.npy", denoised_shots)

    # --- per-shot visualizations ---
    viz_dir = eval_dir / "visualizations"
    viz_dir.mkdir(parents=True, exist_ok=True)
    n_shots = input_shots.shape[0]
    indices = select_random_shots(n_shots, n_viz_shots, seed=42)
    vmax = float(np.quantile(np.abs(np.concatenate([
        input_shots.ravel(), denoised_shots.ravel(), clean_shots.ravel()
    ])), 0.995))

    level_tag = data_level.replace(".", "_")
    title_prefix = f"denoise_{model_type}_level{level_tag}"
    save_shot_visualizations(
        input_shots=input_shots,
        pred_shots=denoised_shots,
        target_shots=clean_shots,
        indices=indices,
        save_dir=viz_dir,
        title_prefix=title_prefix,
        vmin=-vmax,
        vmax=vmax,
        save_npy=True,
    )


def discover_results(root: Path) -> List[Dict[str, Any]]:
    """Scan *root* for experiment directories and return metadata dicts.

    Each dict contains: ``dir`` (Path), ``model``, ``level``, ``seed``.
    Only directories with both ``checkpoints/best.pt`` and ``test_set/``
    are kept.
    """
    entries: List[Dict[str, Any]] = []
    for d in sorted(root.iterdir()):
        if not d.is_dir():
            continue
        parsed = parse_dir_name(d.name)
        if parsed is None:
            # Try simplified format (new training scripts): denoise_{model}_base{date}
            m = _DIR_RE_SIMPLE.match(d.name)
            if m is None:
                print(f"[SKIP] cannot parse directory name: {d.name}")
                continue
            model = m.group(1)
            # Read seed and data level from config.yaml
            seed = "0"
            level = "single"
            config_path = d / "config.yaml"
            if config_path.is_file():
                try:
                    with open(config_path, "r") as f:
                        cfg = yaml.safe_load(f)
                    seed = str(cfg.get("experiment", {}).get("seed", "0"))
                    input_path = cfg.get("data", {}).get("segy_pair", {}).get("input_path", "")
                    lm = _NOISE_LEVEL_RE.search(input_path)
                    if lm:
                        level = lm.group(1)
                except Exception:
                    pass
            parsed = (model, level, seed)
        ckpt = d / "checkpoints" / "best.pt"
        test_dir = d / "test_set"
        if not ckpt.is_file():
            print(f"[SKIP] missing best.pt: {d.name}")
            continue
        if not test_dir.is_dir():
            print(f"[SKIP] missing test_set/: {d.name}")
            continue
        entries.append(
            {
                "dir": d,
                "model": parsed[0],
                "level": parsed[1],
                "seed": parsed[2],
            }
        )
    return entries


def _read_config_norm_mode(result_dir: Path) -> Tuple[str, float, float]:
    """Read normalization settings from experiment config.

    Returns ``(normalize_mode, psnr_peak, ssim_data_range)``.
    For ``max_abs`` data: peak=1.0, range=2.0.
    For ``mean_std`` data: peak/range estimated from actual clean signal.
    """
    config_path = result_dir / "config.yaml"
    normalize_mode = "max_abs"
    if config_path.is_file():
        try:
            with open(config_path, "r") as f:
                cfg = yaml.safe_load(f)
            normalize_mode = str(
                cfg.get("preprocess", {}).get("normalize_mode", "max_abs")
            )
        except Exception:
            pass
    if normalize_mode == "mean_std":
        # Actual peak/range will be estimated from the test data later
        return normalize_mode, -1.0, -1.0
    return normalize_mode, 1.0, 2.0


def _read_config_dt(result_dir: Path) -> float:
    """Read ``dt`` from the experiment config.yaml, defaulting to 0.002 s."""
    config_path = result_dir / "config.yaml"
    if config_path.is_file():
        try:
            with open(config_path, "r") as f:
                cfg = yaml.safe_load(f)
            time_downsample = int(
                cfg.get("data", {})
                .get("segy_pair", {})
                .get("time_downsample", 1)
            )
            return 0.002 * time_downsample
        except Exception:
            pass
    return 0.002


def load_model_from_checkpoint(ckpt_path: Path, device: torch.device) -> torch.nn.Module:
    """Load a model from a training checkpoint."""
    ckpt = torch.load(ckpt_path, map_location=device, weights_only=False)
    model_cfg = ckpt["extras"]["config"]["model"]
    model = build_model(model_cfg)
    model.load_state_dict(ckpt["model"])
    model.to(device)
    model.eval()
    return model


def evaluate_one(
    result_dir: Path,
    device: torch.device,
    patch_size: Tuple[int, int] = (128, 256),
    overlap: float = 0.5,
    batch_size: int = 8,
    n_viz_shots: int = 5,
    save_viz: bool = True,
) -> Optional[Dict[str, Any]]:
    """Run inference + metric computation for a single experiment directory.

    Returns a dict of before/after metrics or ``None`` on failure.
    """
    ckpt_path = result_dir / "checkpoints" / "best.pt"
    test_dir = result_dir / "test_set"

    # --- load model ---------------------------------------------------------
    try:
        model = load_model_from_checkpoint(ckpt_path, device)
        total_params = sum(p.numel() for p in model.parameters())
        num_params_m = total_params / 1e6
        # read model type for visualization prefix
        config_path = result_dir / "config.yaml"
        model_type = "unknown"
        if config_path.is_file():
            try:
                with open(config_path, "r") as f:
                    _cfg = yaml.safe_load(f)
                model_type = str(_cfg.get("model", {}).get("type", "unknown"))
            except Exception:
                pass
    except Exception:
        print(f"[ERROR] failed to load model from {ckpt_path}:")
        traceback.print_exc()
        return None

    # --- load test data -----------------------------------------------------
    try:
        input_shots = np.load(test_dir / "input_shots.npy")
        target_shots = np.load(test_dir / "target_shots.npy")
    except Exception:
        print(f"[ERROR] failed to load test data from {test_dir}:")
        traceback.print_exc()
        return None

    # --- inference ----------------------------------------------------------
    try:
        pred_noise = inference_on_shots(
            model,
            input_shots,
            patch_size=patch_size,
            overlap=overlap,
            device=device,
            batch_size=batch_size,
        )
    except Exception:
        print(f"[ERROR] inference failed for {result_dir.name}:")
        traceback.print_exc()
        return None

    # --- compute signals (3D) ------------------------------------------------
    clean_ref = input_shots - target_shots          # ground-truth clean
    denoised = input_shots - pred_noise             # model output

    # --- save full-volume npy and per-shot visualizations ---------------------
    if save_viz:
        eval_dir = result_dir / "evaluation"
        try:
            noise_level = _parse_noise_level_from_dir(result_dir)
            _save_viz_and_npy(
                eval_dir,
                input_shots,
                clean_ref,
                denoised,
                model_type=model_type,
                data_level=noise_level,
                n_viz_shots=n_viz_shots,
            )
        except Exception:
            print(f"[WARN] visualization saving failed for {result_dir.name}:")
            traceback.print_exc()

    # --- binned metrics (EB-WSE / FB-FRE) on 3D data before flattening ---------
    dt = float(_read_config_dt(result_dir))
    binned_after: Dict[str, Any] = {}
    try:
        binned_after = compute_binned_metrics(denoised, clean_ref, dt=dt)
    except Exception:
        print(f"[WARN] binned metrics failed for {result_dir.name}:")
        traceback.print_exc()

    # --- flatten to 2D: (n_shots × n_traces, n_time) ------------------------
    n_time = input_shots.shape[-1]
    noisy_2d = input_shots.reshape(-1, n_time)
    clean_2d = clean_ref.reshape(-1, n_time)
    denoised_2d = denoised.reshape(-1, n_time)

    # --- read normalization mode for correct PSNR/SSIM data ranges -----------
    norm_mode, psnr_peak, ssim_range = _read_config_norm_mode(result_dir)
    if psnr_peak < 0:
        # mean_std: estimate peak from the actual clean signal
        psnr_peak = float(np.quantile(np.abs(clean_ref), 0.999))
        ssim_range = psnr_peak * 2.0

    # --- metrics (reshape 2D → 3D for compute_shot_metrics) -----------------
    _, before_mean = compute_shot_metrics(
        noisy_2d.reshape(1, -1, n_time), clean_2d.reshape(1, -1, n_time),
        METRIC_NAMES, psnr_peak=psnr_peak, ssim_data_range=ssim_range,
    )
    _, after_mean = compute_shot_metrics(
        denoised_2d.reshape(1, -1, n_time), clean_2d.reshape(1, -1, n_time),
        METRIC_NAMES, psnr_peak=psnr_peak, ssim_data_range=ssim_range,
    )

    return {
        "before": before_mean,
        "after": after_mean,
        "after_binned": binned_after,
        "num_params_m": num_params_m,
    }


def aggregate(
    entries: List[Dict[str, Any]],
) -> Dict[str, Dict[str, Any]]:
    """Aggregate per-seed results into mean±std per (data level, model).

    Returns ``{level: {model: {metric: (mean, std)}, ..., "raw": {metric: val}}}``.
    """
    # group: level → model → list of per-seed metric dicts
    groups: Dict[str, Dict[str, List[Dict[str, float]]]] = defaultdict(
        lambda: defaultdict(list)
    )
    raw_by_level: Dict[str, Dict[str, float]] = {}
    params_by_model: Dict[str, float] = {}

    for entry in entries:
        level = entry["level"]
        model = entry["model"]
        after = entry.get("after", {})
        if after:
            groups[level][model].append(after)
        # capture raw (before) metrics — deterministic, just take the first
        before = entry.get("before", {})
        if before and level not in raw_by_level:
            raw_by_level[level] = before
        # param count — same for all seeds of a model, take first
        n_params = entry.get("num_params_m")
        if n_params is not None and model not in params_by_model:
            params_by_model[model] = n_params

    # discover all binned metric keys present in the results
    binned_keys: set = set()
    for entry in entries:
        ab = entry.get("after_binned", {})
        if ab:
            binned_keys.update(ab.keys())
    binned_keys_sorted = sorted(binned_keys)

    # compute mean±std
    aggregated: Dict[str, Dict[str, Any]] = {}
    for level in sorted(groups.keys(), key=_level_sort_key):
        aggregated[level] = {}
        # raw row
        if level in raw_by_level:
            aggregated[level]["raw"] = raw_by_level[level]
        for model in MODEL_ROW_ORDER:
            seeds = groups[level].get(model, [])
            if not seeds:
                continue
            model_stats: Dict[str, Tuple[float, float]] = {}
            for m in METRIC_NAMES:
                vals = [s[m] for s in seeds if m in s]
                if len(vals) == 0:
                    continue
                mean = float(np.mean(vals))
                if len(vals) >= 2:
                    std = float(np.std(vals, ddof=1))
                else:
                    std = 0.0
                model_stats[m] = (mean, std)
            aggregated[level][model] = model_stats

    # aggregate binned metrics separately: level → model → binned_stats
    binned_groups: Dict[str, Dict[str, List[Dict[str, Any]]]] = defaultdict(
        lambda: defaultdict(list)
    )
    for entry in entries:
        ab = entry.get("after_binned", {})
        if ab:
            binned_groups[entry["level"]][entry["model"]].append(ab)

    binned_aggregated: Dict[str, Dict[str, Dict[str, Tuple[float, float]]]] = {}
    for level in sorted(binned_groups.keys(), key=_level_sort_key):
        binned_aggregated[level] = {}
        for model in MODEL_ROW_ORDER:
            seeds = binned_groups[level].get(model, [])
            if not seeds:
                continue
            model_stats: Dict[str, Tuple[float, float]] = {}
            for k in binned_keys_sorted:
                vals = []
                for s in seeds:
                    v = s.get(k)
                    if v is None:
                        continue
                    if not np.isscalar(v):
                        continue
                    if np.isfinite(v):
                        vals.append(v)
                if len(vals) == 0:
                    continue
                mean = float(np.mean(vals))
                if len(vals) >= 2:
                    std = float(np.std(vals, ddof=1))
                else:
                    std = 0.0
                model_stats[k] = (mean, std)
            if model_stats:
                binned_aggregated[level][model] = model_stats

    return aggregated, params_by_model, binned_keys_sorted, binned_aggregated


def _fmt(mean: float, std: float, metric: str) -> str:
    """Format a cell value as ``mean±std`` with metric-appropriate precision.

    MAE, MSE, RMSE: mean to 6 decimal places, std to 2 decimal places.
    SNR, PSNR, SSIM: mean and std to 2 decimal places.
    """
    if metric in ("mae", "mse", "rmse"):
        return f"{mean:.6f}±{std:.6f}"
    if metric in ("snr", "psnr", "ssim"):
        return f"{mean:.4f}±{std:.4f}"
    return f"{mean:.2f}±{std:.2f}"


def load_existing_results(
    xlsx_path: Path,
) -> Tuple[Dict[str, Dict[str, Any]], Dict[str, float]]:
    """Parse an existing batch evaluation Excel back into aggregated form.

    Returns ``(aggregated, params_by_model)`` with the same structure as
    ``aggregate()``, so new results can be merged in.

    ``aggregated[level][model]`` maps metric name → ``(mean, std)`` tuple.
    For the ``"raw"`` key, the value is a plain ``{metric: float}`` dict.
    """
    try:
        import openpyxl
    except ImportError:
        print("openpyxl is required for --merge.  Install it with:  pip install openpyxl")
        return {}, {}

    # reverse lookup: display name → model key
    display_to_key = {v: k for k, v in MODEL_DISPLAY.items()}
    display_to_key["Raw (noisy)"] = "raw"

    # regex to parse "mean±std" cell values
    _CELL_RE = re.compile(r"^([\d.eE+-]+)±([\d.eE+-]+)$")

    def _parse_cell(val: str):
        """Try to parse 'mean±std' back to (float, float), or None."""
        m = _CELL_RE.match(str(val).strip())
        if m:
            return float(m.group(1)), float(m.group(2))
        return None

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    aggregated: Dict[str, Dict[str, Any]] = {}
    binned_agg: Dict[str, Dict[str, Any]] = {}
    params_by_model: Dict[str, float] = {}

    for sheet_name in wb.sheetnames:
        # sheet name format: "Noise {level}" or "Multiples" for the fixed dataset
        level = "single" if sheet_name == "Multiples" else sheet_name.replace("Noise ", "").strip()
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        # headers: ["method", "parameters (m)", "snr", "psnr", "ssim", "mae", "mse", "rmse"]

        level_data: Dict[str, Any] = {}
        for row in rows[1:]:
            if not row or not row[0]:
                continue
            method_display = str(row[0]).strip()
            model_key = display_to_key.get(method_display)
            if model_key is None:
                continue

            # parse Parameters (M)
            if row[1] is not None and str(row[1]).strip() not in ("—", ""):
                try:
                    params_by_model[model_key] = float(str(row[1]).strip())
                except ValueError:
                    pass

            # parse metrics from columns onward (standard + binned)
            metrics: Dict[str, Any] = {}
            for ci in range(2, len(headers)):
                metric_name = headers[ci]
                is_binned = metric_name.startswith("eb_wse_") or metric_name.startswith("fb_fre_")
                if metric_name not in METRIC_NAMES and not is_binned:
                    continue
                cell_val = row[ci] if ci < len(row) else None
                if cell_val is None or str(cell_val).strip() in ("—", ""):
                    continue
                if model_key == "raw":
                    # just a plain value
                    try:
                        metrics[metric_name] = float(str(cell_val).strip())
                    except ValueError:
                        pass
                else:
                    parsed = _parse_cell(str(cell_val))
                    if parsed is not None:
                        metrics[metric_name] = parsed

            if metrics:
                # split standard vs binned metrics
                std_metrics = {k: v for k, v in metrics.items() if not k.startswith(("eb_wse_", "fb_fre_"))}
                bin_metrics = {k: v for k, v in metrics.items() if k.startswith(("eb_wse_", "fb_fre_"))}
                if std_metrics:
                    level_data.setdefault(model_key, {}).update(std_metrics)
                if bin_metrics:
                    binned_agg.setdefault(level, {}).setdefault(model_key, {}).update(bin_metrics)

        if level_data:
            aggregated[level] = level_data
        if level in binned_agg:
            if level not in binned_agg:
                binned_agg[level] = {}
            binned_agg[level].update(binned_agg.get(level, {}))

    wb.close()
    return aggregated, params_by_model, binned_agg


def build_excel(
    aggregated: Dict[str, Dict[str, Any]],
    params_by_model: Dict[str, float],
    output_path: Path,
    binned_keys: Optional[list] = None,
    binned_aggregated: Optional[Dict[str, Dict[str, Dict[str, Tuple[float, float]]]]] = None,
) -> None:
    """Write one sheet per data level. Rows = methods, columns = metrics + binned."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    except ImportError:
        print(
            "openpyxl is required for Excel output. Install it with:\n"
            "    pip install openpyxl"
        )
        raise

    wb = openpyxl.Workbook()
    # remove default sheet — we'll create one per level
    wb.remove(wb.active)

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    n_metric_cols = len(METRIC_DISPLAY)

    for level in sorted(aggregated.keys(), key=_level_sort_key):
        ws = wb.create_sheet(title=_sheet_title(level))

        # column A = Method, B = Parameters (M), C..H = metrics
        ws.cell(row=1, column=1, value="Method").font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).alignment = center_align
        ws.cell(row=1, column=1).border = thin_border

        cell = ws.cell(row=1, column=2, value="Parameters (M)")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

        for ci, m_disp in enumerate(METRIC_DISPLAY, start=3):
            cell = ws.cell(row=1, column=ci, value=m_disp)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # --- binned metric columns (EB-WSE / FB-FRE) -------------------------
        binned_col_start = 3 + len(METRIC_DISPLAY)
        _binned_keys: list = binned_keys or []
        for bi, bk in enumerate(_binned_keys):
            cell = ws.cell(row=1, column=binned_col_start + bi, value=bk.upper())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        level_data = aggregated[level]
        level_binned = (binned_aggregated or {}).get(level, {})

        # row order: raw, then each model in MODEL_ROW_ORDER
        row_order = ["raw"] + [m for m in MODEL_ROW_ORDER if m in level_data]
        for ri, key in enumerate(row_order, start=2):
            # method label
            label = "Raw (noisy)" if key == "raw" else MODEL_DISPLAY.get(key, key)
            cell = ws.cell(row=ri, column=1, value=label)
            cell.font = Font(bold=True) if key == "raw" else Font()
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border

            # param count
            if key == "raw":
                cell = ws.cell(row=ri, column=2, value="—")
            else:
                n_m = params_by_model.get(key)
                cell = ws.cell(row=ri, column=2,
                               value=round(n_m, 2) if n_m is not None else "—")
            cell.alignment = center_align
            cell.border = thin_border

            data = level_data[key]
            for ci, m_name in enumerate(METRIC_NAMES, start=3):
                if m_name not in data:
                    cell = ws.cell(row=ri, column=ci, value="—")
                elif key == "raw":
                    val = data[m_name]
                    cell = ws.cell(row=ri, column=ci, value=_fmt(val, 0.0, m_name))
                else:
                    mean, std = data[m_name]
                    cell = ws.cell(row=ri, column=ci, value=_fmt(mean, std, m_name))
                cell.alignment = center_align
                cell.border = thin_border

            # --- binned metric cells ---
            binned_data = level_binned.get(key, {}) if key != "raw" else {}
            for bi, bk in enumerate(_binned_keys):
                bd = binned_data.get(bk)
                if bd is not None:
                    val = _fmt(bd[0], bd[1], bk)
                else:
                    val = "—"
                cell = ws.cell(row=ri, column=binned_col_start + bi, value=val)
                cell.alignment = center_align
                cell.border = thin_border

        # auto-width
        from openpyxl.utils import get_column_letter
        total_cols = binned_col_start + len(_binned_keys)
        for ci in range(1, total_cols):
            max_w = 0
            for row in ws.iter_rows(min_col=ci, max_col=ci):
                for c in row:
                    if c.value is not None:
                        max_w = max(max_w, len(str(c.value)))
            ws.column_dimensions[get_column_letter(ci)].width = max_w + 4

        ws.freeze_panes = "C2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Saved Excel to: {output_path}")


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------
def main() -> None:
    parser = argparse.ArgumentParser(
        description="Batch evaluate denoising results on held-out test sets"
    )
    parser.add_argument(
        "--root_dir", default="/data/shared/benchmark/multiples/results", type=Path,
        help="Directory containing experiment subdirectories",
    )
    parser.add_argument(
        "--output", default="scripts/multiples_attenuation/batch_evaluation_part.xlsx", type=Path,
        help="Path for the output Excel file (.xlsx)",
    )
    parser.add_argument(
        "--device", default="cuda:7" if torch.cuda.is_available() else "cpu",
        help="Device for inference (default: cuda if available, else cpu)",
    )
    parser.add_argument(
        "--batch_size", type=int, default=8,
        help="Batch size for inference (default: 8)",
    )
    parser.add_argument(
        "--n-viz-shots",
        type=int,
        default=5,
        help="Number of random shots to visualize per experiment (default: 5).",
    )
    parser.add_argument(
        "--save-viz",
        dest="save_viz",
        action="store_true",
        default=True,
        help="Save full-volume .npy files and per-shot visualizations (default: True).",
    )
    parser.add_argument(
        "--no-save-viz",
        dest="save_viz",
        action="store_false",
        help="Disable saving .npy files and per-shot visualizations.",
    )
    parser.add_argument(
        "--models", nargs="+",
        choices=list(MODEL_DISPLAY.keys()),
        default=None,
        help="Only evaluate the specified model(s).  Choices: %(choices)s.  "
             "If omitted, all discovered models are evaluated.",
    )
    parser.add_argument(
        "--merge", action="store_true",
        help="If the output file already exists, load existing results, "
             "skip already-evaluated (data-level, model) pairs, and merge new "
             "results into the same Excel file without overwriting old data.",
    )
    args = parser.parse_args()

    root_dir: Path = args.root_dir
    if not root_dir.is_dir():
        print(f"Root directory does not exist: {root_dir}")
        sys.exit(1)

    device = torch.device(args.device)
    print(f"Using device: {device}")

    # --- discover -----------------------------------------------------------
    entries = discover_results(root_dir)
    if not entries:
        print("No valid experiment directories found.")
        sys.exit(0)

    print(f"Found {len(entries)} experiment(s) to evaluate.")

    # --- filter by model ------------------------------------------------------
    if args.models:
        entries = [e for e in entries if e["model"] in args.models]
        if not entries:
            print(f"No experiments match the specified model(s): {args.models}")
            sys.exit(0)
        print(f"Filtered to {len(entries)} experiment(s) for: {', '.join(args.models)}")

    # --- merge mode: skip already-evaluated (data-level, model) pairs -----------
    old_aggregated: Dict[str, Dict[str, Any]] = {}
    old_binned: Dict[str, Dict[str, Any]] = {}
    old_params: Dict[str, float] = {}
    if args.merge and args.output.is_file():
        print(f"Merge mode: loading existing results from {args.output}")
        old_aggregated, old_params, old_binned = load_existing_results(args.output)
        # Check whether old Excel already has binned columns
        _old_has_binned = any(
            any(k.startswith("eb_wse_") or k.startswith("fb_fre_") for k in stats)
            for models in old_aggregated.values()
            for stats in models.values()
        )
        # build set of (data-level, model) already in the existing Excel
        already_done: set = set()
        for level, models in old_aggregated.items():
            for model in models:
                # Only skip if old Excel already has binned metrics for this model;
                # otherwise re-evaluate to fill in missing binned columns.
                if _old_has_binned:
                    already_done.add((level, model))
        n_before = len(entries)
        if _old_has_binned:
            entries = [
                e for e in entries
                if (e["level"], e["model"]) not in already_done
            ]
        skipped = n_before - len(entries)
        if skipped:
            skipped_names = sorted(
                {e["model"] for e in [
                    {"model": m, "level": l} for l, m in already_done
                ]}
            )
            print(f"Skipping {skipped} already-evaluated entry(s) for: "
                  f"{', '.join(MODEL_DISPLAY.get(m, m) for m in skipped_names)}")
        if not entries and _old_has_binned:
            print("All requested models already evaluated — nothing to do.")
            sys.exit(0)

    # --- evaluate -----------------------------------------------------------
    for i, entry in enumerate(entries):
        d = entry["dir"]
        print(
            f"[{i + 1}/{len(entries)}] {d.name}  "
            f"(model={entry['model']}, level={entry['level']}, seed={entry['seed']})"
        )
        result = evaluate_one(
            d,
            device=device,
            batch_size=args.batch_size,
            n_viz_shots=args.n_viz_shots,
            save_viz=args.save_viz,
        )
        if result is None:
            print(f"  -> FAILED, skipping")
            entry["before"] = {}
            entry["after"] = {}
        else:
            entry["before"] = result["before"]
            entry["after"] = result["after"]
            entry["after_binned"] = result.get("after_binned", {})
            entry["num_params_m"] = result["num_params_m"]
            b = result["before"]
            a = result["after"]
            print(
                f"  Params: {result['num_params_m']:.2f}M  |  "
                f"SNR:  {b['snr']:>7.4f} -> {a['snr']:>7.4f} dB  |  "
                f"PSNR: {b['psnr']:>7.4f} -> {a['psnr']:>7.4f} dB  |  "
                f"SSIM: {b['ssim']:.4f} -> {a['ssim']:.4f}  |  "
                f"MSE:  {b['mse']:.6f} -> {a['mse']:.6f}"
            )
            # Print binned metric summary
            ab = result.get("after_binned", {})
            if ab:
                snr_items = {k: v for k, v in ab.items() if "snr" in k}
                if snr_items:
                    parts = [f"{k}={v:.2f}" if v is not None else f"{k}=N/A" for k, v in sorted(snr_items.items())]
                    print(f"  Binned SNR: {', '.join(parts)}")

    # --- aggregate per level -------------------------------------------------
    aggregated, params_by_model, binned_keys_sorted, binned_aggregated = aggregate(entries)

    # --- merge with old results if requested ----------------------------------
    if old_aggregated:
        for level, models in old_aggregated.items():
            if level not in aggregated:
                aggregated[level] = {}
            for model, stats in models.items():
                if model not in aggregated[level]:
                    aggregated[level][model] = stats
        for model, n_m in old_params.items():
            if model not in params_by_model:
                params_by_model[model] = n_m
        # re-sort data levels
        aggregated = {k: aggregated[k] for k in sorted(aggregated.keys(), key=_level_sort_key)}
    # merge old binned data
    if old_binned:
        for level, models in old_binned.items():
            binned_aggregated.setdefault(level, {})
            for model, stats in models.items():
                if model not in binned_aggregated[level]:
                    binned_aggregated[level][model] = stats
        # discover binned keys from old data too
        for models in old_binned.values():
            for stats in models.values():
                for k in stats:
                    if k not in binned_keys_sorted:
                        binned_keys_sorted.append(k)
        binned_keys_sorted = sorted(binned_keys_sorted)

    # summary of what was aggregated
    print("Model parameter counts:")
    for m in MODEL_ROW_ORDER:
        if m in params_by_model:
            print(f"  {MODEL_DISPLAY.get(m, m):<16s}  {params_by_model[m]:.2f}M")
    for level in sorted(aggregated.keys(), key=_level_sort_key):
        models = [k for k in aggregated[level] if k != "raw"]
        label = "Multiples" if level == "single" else f"Noise {level}"
        print(f"{label}: raw + {len(models)} model(s) — {', '.join(MODEL_DISPLAY.get(m, m) for m in models)}")

    # --- export -------------------------------------------------------------
    build_excel(aggregated, params_by_model, args.output,
                binned_keys=binned_keys_sorted, binned_aggregated=binned_aggregated)


if __name__ == "__main__":
    main()
